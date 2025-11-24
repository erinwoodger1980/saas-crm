#!/usr/bin/env python3
"""
Extract door specifications, dropdowns, and formulas from Excel file
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json
import re

# Load the workbook
wb = openpyxl.load_workbook('/Users/Erin/Desktop/open ai quote sheet no macros.xlsx', data_only=False)
ws = wb.active

print("=" * 80)
print(f"Worksheet: {ws.title}")
print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
print("=" * 80)

# Find header row (look for common door spec headers)
header_row = None
for row_idx in range(1, min(50, ws.max_row + 1)):
    row_values = [str(cell.value).lower() if cell.value else '' for cell in ws[row_idx]]
    if any(h in ' '.join(row_values) for h in ['door', 'qty', 'width', 'height', 'location']):
        header_row = row_idx
        print(f"\nFound header row at: {row_idx}")
        break

if header_row:
    # Extract headers
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col_idx)
        if cell.value:
            headers.append({
                'col': get_column_letter(col_idx),
                'index': col_idx,
                'name': str(cell.value).strip(),
                'fill': str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor else None
            })
    
    print(f"\n{'=' * 80}")
    print(f"HEADERS ({len(headers)} columns):")
    print(f"{'=' * 80}")
    for h in headers:
        print(f"  {h['col']:3s} (#{h['index']:3d}): {h['name'][:50]:50s} [color: {h['fill']}]")
    
    # Look for data validation (dropdowns)
    print(f"\n{'=' * 80}")
    print("DROPDOWNS/DATA VALIDATION:")
    print(f"{'=' * 80}")
    
    dropdown_fields = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type == 'list':
            formula = dv.formula1
            cells = str(dv.sqref)
            print(f"\n  Range: {cells}")
            print(f"  Formula: {formula}")
            
            # Try to extract list values
            if formula:
                if formula.startswith('"'):
                    # Inline list
                    values = [v.strip() for v in formula.strip('"').split(',')]
                    print(f"  Values: {values}")
                    dropdown_fields[cells] = {'type': 'inline', 'values': values}
                elif formula.startswith('$') or '!' in formula:
                    # Reference to range
                    print(f"  Reference: {formula}")
                    dropdown_fields[cells] = {'type': 'reference', 'formula': formula}
    
    # Sample first few data rows
    print(f"\n{'=' * 80}")
    print("SAMPLE DATA ROWS (first 5 after header):")
    print(f"{'=' * 80}")
    
    for row_idx in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        print(f"\nRow {row_idx}:")
        for h in headers[:20]:  # First 20 columns
            cell = ws.cell(row_idx, h['index'])
            value = cell.value
            formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
            if hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('='):
                formula = cell._value
            
            display = str(value)[:40] if value else ''
            if formula:
                print(f"  {h['name'][:30]:30s}: {display:40s} [FORMULA: {formula[:60]}]")
            elif value:
                print(f"  {h['name'][:30]:30s}: {display}")
    
    # Look for formulas in first data column
    print(f"\n{'=' * 80}")
    print("FORMULAS FOUND (sampling columns):")
    print(f"{'=' * 80}")
    
    formula_samples = {}
    for row_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
        for h in headers:
            cell = ws.cell(row_idx, h['index'])
            # Check if cell has a formula
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                if h['name'] not in formula_samples:
                    formula_samples[h['name']] = []
                formula_samples[h['name']].append(cell.value)
    
    for field_name, formulas in formula_samples.items():
        if formulas:
            print(f"\n  {field_name}:")
            for f in formulas[:3]:  # Show first 3 examples
                print(f"    {f}")

    # Export to JSON
    output = {
        'sheet_name': ws.title,
        'header_row': header_row,
        'headers': headers,
        'dropdowns': dropdown_fields,
        'formula_samples': formula_samples
    }
    
    with open('/Users/Erin/saas-crm/door-spec-structure.json', 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"\n{'=' * 80}")
    print("Exported structure to: door-spec-structure.json")
    print(f"{'=' * 80}")

else:
    print("Could not find header row!")

wb.close()
