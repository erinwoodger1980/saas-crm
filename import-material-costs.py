#!/usr/bin/env python3
"""
Import ALL material costs from Door Production spreadsheet into database.
Populates DoorCore, IronmongeryItem, and MaterialItem tables with costs.
Ensures door costing matches spreadsheet exactly.
"""
import openpyxl
import psycopg2
from decimal import Decimal
import os
import sys
import re
import uuid

def load_env():
    """Load .env file from current or parent directory if DATABASE_URL not already set"""
    if 'DATABASE_URL' in os.environ:
        return
    candidate_paths = [os.getcwd(), os.path.dirname(__file__), os.path.join(os.getcwd(), 'api')]
    for path in candidate_paths:
        env_path = os.path.join(path, '.env')
        if os.path.exists(env_path):
            try:
                with open(env_path) as f:
                    for line in f:
                        m = re.match(r'([A-Za-z_][A-Za-z0-9_]*)=(.*)', line.strip())
                        if m:
                            key, val = m.groups()
                            if key not in os.environ:
                                os.environ[key] = val
                break
            except Exception as e:
                print(f"Warning: failed to load env file {env_path}: {e}")

load_env()

# Database configuration (after env load)
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/postgres')

# Tenant ID - REQUIRED parameter
TENANT_ID = sys.argv[1] if len(sys.argv) > 1 else None
if not TENANT_ID:
    print("ERROR: Tenant ID required")
    print("Usage: python import-material-costs.py <tenant_id>")
    sys.exit(1)

EXCEL_PATH = '/Users/Erin/Desktop/open ai quote sheet no macros.xlsx'

def connect_db():
    """Connect to PostgreSQL database"""
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def find_sheet(wb, keywords):
    """Find sheet by keywords in name"""
    for name in wb.sheetnames:
        if all(kw.lower() in name.lower() for kw in keywords):
            return name
    return None

def parse_core_materials(ws):
    """Parse door core materials from Door Production sheet"""
    print("\n" + "="*80)
    print("PARSING DOOR CORES")
    print("="*80)
    
    # Look for core types in dropdown lists or reference sheets
    # These are typically on separate sheets or in named ranges
    cores = []
    
    # Common core specifications from the Door Production structure
    # We'll extract unique values from the CORE and CORE TYPE columns
    core_col = None
    rating_col = None
    
    # Find header row
    for row_idx in range(1, 20):
        for col_idx in range(1, 50):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and str(cell.value).strip().upper() == 'CORE':
                core_col = col_idx
            if cell.value and str(cell.value).strip().upper() == 'RATING':
                rating_col = col_idx
        if core_col and rating_col:
            header_row = row_idx
            break
    
    if not core_col:
        print("WARNING: Could not find CORE column")
        return cores
    
    print(f"Found CORE column at index {core_col}, RATING at {rating_col}")
    
    # Extract unique core values from data rows
    seen_cores = set()
    for row_idx in range(header_row + 1, min(header_row + 1000, ws.max_row + 1)):
        core_cell = ws.cell(row_idx, core_col)
        rating_cell = ws.cell(row_idx, rating_col) if rating_col else None
        
        if core_cell.value:
            core_name = str(core_cell.value).strip()
            rating = str(rating_cell.value).strip() if rating_cell and rating_cell.value else 'FD30'
            
            key = (core_name, rating)
            if key not in seen_cores and core_name:
                seen_cores.add(key)
                cores.append({
                    'code': core_name.upper().replace(' ', '_')[:50],
                    'name': core_name,
                    'supplier': 'TBC',
                    'unitCost': Decimal('0.00'),  # Will need manual update
                    'currency': 'GBP',
                    'fireRating': rating,
                    'maxHeight': Decimal('2400'),
                    'maxWidth': Decimal('1200'),
                    'isActive': True
                })
    
    print(f"Extracted {len(cores)} unique door cores")
    for core in cores[:10]:
        print(f"  - {core['name']} ({core['fireRating']})")
    
    return cores

def parse_core_prices_sheet(wb):
    """Parse detailed core prices from 'Door Core Prices' sheet if present"""
    sheet_name = None
    for name in wb.sheetnames:
        if 'core' in name.lower() and 'price' in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        return []
    print("\n" + "="*80)
    print("PARSING DOOR CORE PRICES")
    print("="*80)
    ws = wb[sheet_name]
    # Identify header row (look for 'Product' and 'Price')
    header_row = None
    headers = {}
    for r in range(1, min(10, ws.max_row + 1)):
        present = 0
        local_headers = {}
        for c in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if val:
                name = str(val).strip().lower()
                local_headers[name] = c
                if name in ['product', 'price', 'ogl code', 'thickness (mm)', 'height (mm)', 'width (mm)', 'type']:
                    present += 1
        if 'product' in local_headers and 'price' in local_headers:
            header_row = r
            headers = local_headers
            break
    if not header_row:
        print("No header row found in core prices sheet")
        return []
    print(f"Core prices header row: {header_row}")
    price_col = headers.get('price')
    product_col = headers.get('product')
    code_col = headers.get('ogl code')
    thick_col = headers.get('thickness (mm)')
    height_col = headers.get('height (mm)')
    width_col = headers.get('width (mm)')
    type_col = headers.get('type')
    cores = []
    seen_codes = set()
    for r in range(header_row + 1, ws.max_row + 1):
        product = ws.cell(r, product_col).value if product_col else None
        price = ws.cell(r, price_col).value if price_col else None
        code = ws.cell(r, code_col).value if code_col else None
        thickness = ws.cell(r, thick_col).value if thick_col else None
        h_val = ws.cell(r, height_col).value if height_col else None
        w_val = ws.cell(r, width_col).value if width_col else None
        type_val = ws.cell(r, type_col).value if type_col else None
        if not product or not price:
            continue
        try:
            price_dec = Decimal(str(price))
        except Exception:
            continue
        code_str = str(code).strip() if code else f"CORE_{str(product).strip().upper()}_{r}" if product else f"CORE_{r}"
        code_norm = re.sub(r'[^A-Z0-9_.]', '_', code_str.upper())[:50]
        if code_norm in seen_codes:
            continue
        seen_codes.add(code_norm)
        thickness_num = None
        try:
            thickness_num = float(thickness) if thickness is not None else None
        except Exception:
            thickness_num = None
        fire_rating = 'FD60' if thickness_num and thickness_num >= 54 else 'FD30'
        max_height = Decimal(str(h_val)) if h_val else Decimal('2400')
        max_width = Decimal(str(w_val)) if w_val else Decimal('1200')
        cores.append({
            'code': code_norm,
            'name': f"{product} {type_val or ''}".strip(),
            'supplier': 'Strebord' if product and 'strebord' in str(product).lower() else 'TBC',
            'unitCost': price_dec,
            'currency': 'GBP',
            'fireRating': fire_rating,
            'maxHeight': max_height,
            'maxWidth': max_width,
            'isActive': True
        })
    print(f"Extracted {len(cores)} priced cores from '{sheet_name}'")
    for c in cores[:10]:
        print(f"  - {c['code']} £{c['unitCost']} {c['fireRating']}")
    return cores

def parse_ironmongery(ws):
    """Parse ironmongery items (hinges, locks, handles) from Door Production sheet"""
    print("\n" + "="*80)
    print("PARSING IRONMONGERY")
    print("="*80)
    
    items = []
    
    # Find columns for hinges, locks, handles
    header_row = None
    hinge_col = None
    lock_col = None
    
    for row_idx in range(1, 20):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                val = str(cell.value).strip().upper()
                if val in ['HINGE', 'HINGES']:
                    hinge_col = col_idx
                    header_row = row_idx
                elif val == 'LOCK':
                    lock_col = col_idx
                    header_row = row_idx
        if hinge_col or lock_col:
            break
    
    if not header_row:
        print("WARNING: Could not find ironmongery columns")
        return items
    
    print(f"Found HINGE column: {hinge_col}, LOCK column: {lock_col}")
    
    # Extract unique hinge types
    seen_hinges = set()
    if hinge_col:
        for row_idx in range(header_row + 1, min(header_row + 1000, ws.max_row + 1)):
            cell = ws.cell(row_idx, hinge_col)
            if cell.value:
                hinge_name = str(cell.value).strip()
                if hinge_name and hinge_name not in seen_hinges:
                    seen_hinges.add(hinge_name)
                    items.append({
                        'category': 'HINGE',
                        'code': hinge_name.upper().replace(' ', '_')[:50],
                        'name': hinge_name,
                        'supplier': 'TBC',
                        'unitCost': Decimal('0.00'),
                        'currency': 'GBP',
                        'isActive': True
                    })
    
    # Extract unique lock types
    seen_locks = set()
    if lock_col:
        for row_idx in range(header_row + 1, min(header_row + 1000, ws.max_row + 1)):
            cell = ws.cell(row_idx, lock_col)
            if cell.value:
                lock_name = str(cell.value).strip()
                if lock_name and lock_name not in seen_locks:
                    seen_locks.add(lock_name)
                    items.append({
                        'category': 'LOCK',
                        'code': lock_name.upper().replace(' ', '_')[:50],
                        'name': lock_name,
                        'supplier': 'TBC',
                        'unitCost': Decimal('0.00'),
                        'currency': 'GBP',
                        'isActive': True
                    })
    
    print(f"Extracted {len(items)} ironmongery items")
    print(f"  Hinges: {len([i for i in items if i['category'] == 'HINGE'])}")
    print(f"  Locks: {len([i for i in items if i['category'] == 'LOCK'])}")
    
    return items

def parse_ironmongery_sheet(wb):
    """Parse the dedicated Ironmongery sheet for all items with prices"""
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == 'ironmongery' or 'ironmongery' in name.lower():
            sheet_name = name
            break
    
    if not sheet_name:
        print("\nWARNING: No Ironmongery sheet found")
        return []
    
    print("\n" + "="*80)
    print(f"PARSING IRONMONGERY SHEET: {sheet_name}")
    print("="*80)
    
    ws = wb[sheet_name]
    items = []
    
    # Find header row
    header_row = None
    headers = {}
    for r in range(1, min(20, ws.max_row + 1)):
        row_obj = ws[r]
        for c_idx, cell in enumerate(row_obj, start=1):
            if cell.value:
                h_name = str(cell.value).strip().lower()
                headers[h_name] = c_idx
                if any(kw in h_name for kw in ['description', 'item', 'product', 'code', 'price', 'cost']):
                    header_row = r
    
    if not header_row:
        print("Could not find header row in Ironmongery sheet")
        return []
    
    # Try to find columns
    code_col = headers.get('code') or headers.get('product code') or headers.get('item code')
    desc_col = headers.get('description') or headers.get('item') or headers.get('product')
    price_col = headers.get('price') or headers.get('cost') or headers.get('unit price')
    cat_col = headers.get('category') or headers.get('type')
    
    if not desc_col:
        print("Could not identify description column")
        return []
    
    print(f"Found columns - Description: {desc_col}, Price: {price_col}, Code: {code_col}, Category: {cat_col}")
    
    seen_items = set()
    for r in range(header_row + 1, ws.max_row + 1):
        row_obj = ws[r]
        desc = row_obj[desc_col - 1].value if desc_col else None
        price = row_obj[price_col - 1].value if price_col else None
        code = row_obj[code_col - 1].value if code_col else None
        category = row_obj[cat_col - 1].value if cat_col else None
        
        if not desc:
            continue
        
        desc_str = str(desc).strip()
        if not desc_str or desc_str in seen_items:
            continue
        
        seen_items.add(desc_str)
        
        # Determine category from description or category column
        cat = 'LOCK'
        if category:
            cat_str = str(category).upper()
            if 'HINGE' in cat_str or 'PIVOT' in cat_str:
                cat = 'HINGE'
            elif 'HANDLE' in cat_str or 'LEVER' in cat_str:
                cat = 'LEVER_HANDLE'
            elif 'LOCK' in cat_str or 'LATCH' in cat_str:
                cat = 'LOCK'
        elif 'hinge' in desc_str.lower() or 'pivot' in desc_str.lower():
            cat = 'HINGE'
        elif 'handle' in desc_str.lower() or 'lever' in desc_str.lower():
            cat = 'LEVER_HANDLE'
        
        price_dec = Decimal('0.00')
        if price:
            try:
                price_dec = Decimal(str(price))
            except:
                pass
        
        code_str = str(code).strip() if code else desc_str
        code_norm = re.sub(r'[^A-Z0-9_.]', '_', code_str.upper())[:50]
        
        items.append({
            'category': cat,
            'code': code_norm,
            'name': desc_str,
            'supplier': 'TBC',
            'unitCost': price_dec,
            'currency': 'GBP',
            'isActive': True
        })
    
    print(f"Extracted {len(items)} ironmongery items from sheet")
    print(f"  Hinges: {len([i for i in items if i['category'] == 'HINGE'])}")
    print(f"  Locks: {len([i for i in items if i['category'] == 'LOCK'])}")
    print(f"  Handles: {len([i for i in items if i['category'] == 'LEVER_HANDLE'])}")
    
    return items

def parse_timber_prices(wb):
    """Parse Timber Prices sheet for lippings and linings"""
    sheet_name = None
    for name in wb.sheetnames:
        if 'timber' in name.lower() and 'price' in name.lower():
            sheet_name = name
            break
    
    if not sheet_name:
        print("\nWARNING: No Timber Prices sheet found")
        return []
    
    print("\n" + "="*80)
    print(f"PARSING TIMBER PRICES SHEET: {sheet_name}")
    print("="*80)
    
    ws = wb[sheet_name]
    materials = []
    
    # This sheet has sections like "Lippings" and "Linings"
    current_category = None
    header_row = None
    headers = {}
    
    for r in range(1, min(ws.max_row + 1, 100)):
        first_cell = ws.cell(r, 1).value
        if first_cell:
            first_str = str(first_cell).strip()
            
            # Check if this is a category header
            if first_str.lower() in ['lippings', 'linings', 'boards']:
                current_category = first_str.upper()
                print(f"\nFound category: {current_category}")
                continue
            
            # Check if this is a data header row
            if 'thickness' in first_str.lower() or 'width' in first_str.lower():
                header_row = r
                headers = {}
                for c in range(1, min(15, ws.max_column + 1)):
                    val = ws.cell(r, c).value
                    if val:
                        h_name = str(val).strip().lower()
                        headers[h_name] = c
                print(f"  Header row at {r}: {list(headers.keys())}")
                continue
        
        # Parse data rows
        if current_category and header_row and r > header_row:
            thick_col = headers.get('thickness')
            width_col = headers.get('width')
            length_col = headers.get('length')
            
            if not (thick_col and width_col):
                continue
            
            thickness = ws.cell(r, thick_col).value
            width = ws.cell(r, width_col).value
            length = ws.cell(r, length_col).value if length_col else 2400
            
            if not thickness or not width:
                continue
            
            # Get prices for different timber types
            for h_name, col in headers.items():
                if h_name in ['thickness', 'width', 'length']:
                    continue
                
                price = ws.cell(r, col).value
                if not price:
                    continue
                
                try:
                    price_dec = Decimal(str(price))
                except:
                    continue
                
                # Create material item
                code = f"{current_category}_{h_name}_{thickness}x{width}x{length}".upper().replace(' ', '_')[:50]
                code = re.sub(r'[^A-Z0-9_.]', '_', code)
                
                materials.append({
                    'category': 'LIPPING' if current_category == 'LIPPINGS' else 'TIMBER',
                    'code': code,
                    'name': f"{current_category} {h_name.upper()} {thickness}x{width}x{length}mm",
                    'cost': price_dec,
                    'unit': 'each',
                    'description': f"{thickness}mm thick, {width}mm wide, {length}mm long"
                })
    
    print(f"\nExtracted {len(materials)} timber items")
    for cat in ['LIPPING', 'TIMBER']:
        count = len([m for m in materials if m['category'] == cat])
        print(f"  {cat}: {count}")
    
    return materials

def parse_glass_prices(wb):
    """Parse Glass Prices sheet"""
    sheet_name = None
    for name in wb.sheetnames:
        if 'glass' in name.lower() and 'price' in name.lower():
            sheet_name = name
            break
    
    if not sheet_name:
        print("\nWARNING: No Glass Prices sheet found")
        return []
    
    print("\n" + "="*80)
    print(f"PARSING GLASS PRICES SHEET: {sheet_name}")
    print("="*80)
    
    ws = wb[sheet_name]
    materials = []
    
    # Find header row
    header_row = None
    headers = {}
    for r in range(1, min(20, ws.max_row + 1)):
        present = 0
        local_headers = {}
        for c in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if val:
                h_name = str(val).strip().lower()
                local_headers[h_name] = c
                if any(kw in h_name for kw in ['product', 'glass', 'price', 'rating', 'thickness']):
                    present += 1
        if present >= 3:
            header_row = r
            headers = local_headers
            break
    
    if not header_row:
        print("Could not find header row")
        return []
    
    print(f"Found headers: {list(headers.keys())}")
    
    product_col = headers.get('glass product') or headers.get('product')
    price_col = headers.get('list price per m sq') or headers.get('price') or headers.get('cost')
    rating_col = headers.get('fire rating integrity / insula') or headers.get('fire rating') or headers.get('rating')
    thick_col = headers.get('thickness (mm)') or headers.get('thickness')
    
    if not (product_col and price_col):
        print("Could not identify required columns")
        return []
    
    for r in range(header_row + 1, ws.max_row + 1):
        product = ws.cell(r, product_col).value
        price = ws.cell(r, price_col).value
        rating = ws.cell(r, rating_col).value if rating_col else ''
        thickness = ws.cell(r, thick_col).value if thick_col else ''
        
        if not product or not price:
            continue
        
        product_str = str(product).strip()
        if not product_str or product_str.lower() in ['solid panel', 'vp perspex']:
            continue
        
        try:
            price_dec = Decimal(str(price))
        except:
            continue
        
        code = re.sub(r'[^A-Z0-9_.]', '_', product_str.upper())[:50]
        
        desc_parts = [product_str]
        if rating:
            desc_parts.append(f"Rating: {rating}")
        if thickness:
            desc_parts.append(f"{thickness}mm")
        
        materials.append({
            'category': 'GLASS',
            'code': code,
            'name': product_str,
            'cost': price_dec,
            'unit': 'm2',
            'description': ' | '.join(desc_parts)
        })
    
    print(f"Extracted {len(materials)} glass items")
    return materials

def parse_veneer_prices(wb):
    """Parse Veneer Layon Prices sheet"""
    sheet_name = None
    for name in wb.sheetnames:
        if 'veneer' in name.lower() and 'price' in name.lower():
            sheet_name = name
            break
    
    if not sheet_name:
        print("\nWARNING: No Veneer Prices sheet found")
        return []
    
    print("\n" + "="*80)
    print(f"PARSING VENEER PRICES SHEET: {sheet_name}")
    print("="*80)
    
    ws = wb[sheet_name]
    materials = []
    
    # Find header row (look for SPECIALS or similar)
    header_row = None
    headers = {}
    for r in range(1, min(30, ws.max_row + 1)):
        row_values = []
        for c in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if val:
                val_str = str(val).strip().lower()
                row_values.append(val_str)
                if any(kw in val_str for kw in ['£/m2', 'price', 'vertical', 'horizontal']):
                    headers[val_str] = c
        
        if '£/m2' in ' '.join(row_values) or 'price' in ' '.join(row_values):
            header_row = r
            break
    
    if not header_row:
        print("Could not find header row")
        return []
    
    print(f"Found header row at {header_row}")
    
    # First column should be veneer type
    for r in range(header_row + 1, ws.max_row + 1):
        veneer_name = ws.cell(r, 1).value
        if not veneer_name:
            continue
        
        veneer_str = str(veneer_name).strip()
        if not veneer_str or len(veneer_str) < 2:
            continue
        
        # Get price from second column (usually £/m2 vertical)
        price_cell = ws.cell(r, 2).value
        if not price_cell:
            continue
        
        try:
            price_dec = Decimal(str(price_cell))
        except:
            continue
        
        code = re.sub(r'[^A-Z0-9_.]', '_', f"VENEER_{veneer_str}".upper())[:50]
        
        materials.append({
            'category': 'VENEER',
            'code': code,
            'name': f"{veneer_str} Veneer",
            'cost': price_dec,
            'unit': 'm2',
            'description': f"{veneer_str} veneer layon"
        })
    
    print(f"Extracted {len(materials)} veneer items")
    return materials

def check_lookup_sheets(wb):
    """Check for separate lookup/reference sheets with material costs"""
    print("\n" + "="*80)
    print("CHECKING FOR LOOKUP SHEETS")
    print("="*80)
    
    for sheet_name in wb.sheetnames:
        print(f"\nSheet: {sheet_name}")
    
    return True

def import_door_cores(conn, cores):
    """Insert door cores into database"""
    cursor = conn.cursor()
    
    print(f"\nImporting {len(cores)} door cores...")
    
    for core in cores:
        try:
            core_id = str(uuid.uuid4())
            cursor.execute("""
                INSERT INTO "DoorCore" (
                    id, "tenantId", code, name, supplier, "unitCost", currency,
                    "fireRating", "maxHeight", "maxWidth", "isActive",
                    "createdAt", "updatedAt"
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
                )
                ON CONFLICT ("tenantId", code) DO UPDATE SET
                    name = EXCLUDED.name,
                    supplier = EXCLUDED.supplier,
                    "unitCost" = EXCLUDED."unitCost",
                    "fireRating" = EXCLUDED."fireRating",
                    "updatedAt" = NOW()
            """, (
                core_id,
                TENANT_ID,
                core['code'],
                core['name'],
                core['supplier'],
                core['unitCost'],
                core['currency'],
                core['fireRating'],
                core['maxHeight'],
                core['maxWidth'],
                core['isActive']
            ))
            print(f"  ✓ {core['name']}")
        except Exception as e:
            print(f"  ✗ {core['name']}: {e}")
    
    conn.commit()
    cursor.close()

def import_ironmongery(conn, items):
    """Insert ironmongery items into database"""
    cursor = conn.cursor()
    
    print(f"\nImporting {len(items)} ironmongery items...")
    
    for item in items:
        try:
            item_id = str(uuid.uuid4())
            cursor.execute("""
                INSERT INTO "IronmongeryItem" (
                    id, "tenantId", category, code, name, supplier, "unitCost",
                    currency, "isActive", "createdAt", "updatedAt"
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
                )
                ON CONFLICT ("tenantId", code) DO UPDATE SET
                    category = EXCLUDED.category,
                    name = EXCLUDED.name,
                    supplier = EXCLUDED.supplier,
                    "unitCost" = EXCLUDED."unitCost",
                    "updatedAt" = NOW()
            """, (
                item_id,
                TENANT_ID,
                item['category'],
                item['code'],
                item['name'],
                item['supplier'],
                item['unitCost'],
                item['currency'],
                item['isActive']
            ))
            print(f"  ✓ {item['category']}: {item['name']}")
        except Exception as e:
            print(f"  ✗ {item['name']}: {e}")
    
    conn.commit()
    cursor.close()

def import_material_items(conn, materials):
    """Insert material items (timber, glass, veneer, etc.) into database"""
    cursor = conn.cursor()
    
    print(f"\nImporting {len(materials)} material items...")
    
    for mat in materials:
        try:
            mat_id = str(uuid.uuid4())
            cursor.execute("""
                INSERT INTO "MaterialItem" (
                    id, "tenantId", category, code, name, description, cost,
                    currency, unit, "isActive", "createdAt", "updatedAt"
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
                )
                ON CONFLICT ("tenantId", code) DO UPDATE SET
                    category = EXCLUDED.category,
                    name = EXCLUDED.name,
                    description = EXCLUDED.description,
                    cost = EXCLUDED.cost,
                    unit = EXCLUDED.unit,
                    "updatedAt" = NOW()
            """, (
                mat_id,
                TENANT_ID,
                mat['category'],
                mat['code'],
                mat['name'],
                mat.get('description', ''),
                mat['cost'],
                'GBP',
                mat.get('unit', 'each'),
                True
            ))
            print(f"  ✓ {mat['category']}: {mat['name']} - £{mat['cost']}")
        except Exception as e:
            print(f"  ✗ {mat['name']}: {e}")
    
    conn.commit()
    cursor.close()

def main():
    """Main import process"""
    print("="*80)
    print("MATERIAL COST IMPORT")
    print("="*80)
    print(f"Excel file: {EXCEL_PATH}")
    print(f"Tenant ID: {TENANT_ID}")
    print(f"Database: {DATABASE_URL}")
    
    # Load workbook
    print("\nLoading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    print(f"\nAvailable sheets: {', '.join(wb.sheetnames)}")
    
    # Check for lookup sheets first
    check_lookup_sheets(wb)
    
    # Find Door Production sheet
    door_sheet_name = find_sheet(wb, ['door', 'production'])
    if not door_sheet_name:
        door_sheet_name = wb.sheetnames[0]
    
    print(f"\nUsing sheet: {door_sheet_name}")
    ws = wb[door_sheet_name]
    
    # Parse materials from all sheets
    cores = parse_core_materials(ws)
    
    # Augment with priced cores from dedicated sheet
    priced_cores = parse_core_prices_sheet(wb)
    if priced_cores:
        # Merge by code preference to priced cores
        existing_codes = {c['code']: c for c in cores}
        for pc in priced_cores:
            existing_codes[pc['code']] = pc
        cores = list(existing_codes.values())
    
    # Parse ironmongery from Door Production sheet (legacy)
    ironmongery_old = parse_ironmongery(ws)
    
    # Parse comprehensive data from dedicated sheets
    print("\nParsing dedicated pricing sheets...")
    ironmongery_items = parse_ironmongery_sheet(wb)
    timber_items = parse_timber_prices(wb)
    glass_items = parse_glass_prices(wb)
    veneer_items = parse_veneer_prices(wb)
    
    # Combine all material items (timber, glass, veneer)
    all_material_items = timber_items + glass_items + veneer_items
    
    # Merge ironmongery from both sources (prefer dedicated sheet)
    ironmongery_codes = {item['code']: item for item in ironmongery_old}
    for item in ironmongery_items:
        ironmongery_codes[item['code']] = item
    ironmongery = list(ironmongery_codes.values())
    
    wb.close()
    
    if not cores and not ironmongery and not all_material_items:
        print("\n⚠️  No materials extracted. Please check Excel file structure.")
        sys.exit(1)
    
    # Connect to database
    print("\nConnecting to database...")
    conn = connect_db()
    
    # Import data
    if cores:
        import_door_cores(conn, cores)
    
    if ironmongery:
        import_ironmongery(conn, ironmongery)
    
    if all_material_items:
        import_material_items(conn, all_material_items)
    
    conn.close()
    
    print("\n" + "="*80)
    print("IMPORT COMPLETE")
    print("="*80)
    print(f"Imported {len(cores)} door cores")
    print(f"Imported {len(ironmongery)} ironmongery items")
    print(f"Imported {len(timber_items)} timber items")
    print(f"Imported {len(glass_items)} glass items")
    print(f"Imported {len(veneer_items)} veneer items")
    print(f"Total material items: {len(all_material_items)}")
    print("\n✓ All prices imported from spreadsheet")

if __name__ == '__main__':
    main()
