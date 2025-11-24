#!/usr/bin/env python3
"""
Extract door production sheet structure with actual field names and dropdowns
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

# Load the workbook
wb = openpyxl.load_workbook('/Users/Erin/Desktop/open ai quote sheet no macros.xlsx', data_only=False)

# Find the Door Production sheet
sheet_name = None
for name in wb.sheetnames:
    if 'door' in name.lower() and 'production' in name.lower():
        sheet_name = name
        break

if not sheet_name:
    print("Available sheets:")
    for name in wb.sheetnames:
        print(f"  - {name}")
    print("\nLooking for any sheet with 'door' in name...")
    for name in wb.sheetnames:
        if 'door' in name.lower():
            sheet_name = name
            break

if not sheet_name:
    sheet_name = wb.sheetnames[0]  # Use first sheet

print(f"Using sheet: {sheet_name}")
ws = wb[sheet_name]

print("=" * 80)
print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
print("=" * 80)

# Find headers - look for row with Door Ref, Location, etc.
header_row = None
for row_idx in range(1, min(20, ws.max_row + 1)):
    row_values = []
    for col_idx in range(1, min(30, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            row_values.append(str(cell.value).strip().lower())
   
    # Look for common door specification headers
    if any(kw in ' '.join(row_values) for kw in ['door ref', 'location', 'fire rating', 'qty', 'height', 'width']):
        header_row = row_idx
        print(f"\nFound header row at: {row_idx}")
        print(f"Sample headers: {row_values[:10]}")
        break

if header_row:
    # Extract all headers
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col_idx)
        if cell.value:
            header_name = str(cell.value).strip()
            headers.append({
                'col': get_column_letter(col_idx),
                'index': col_idx,
                'name': header_name,
                'fill': str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor else None
            })
    
    print(f"\n{'=' * 80}")
    print(f"ALL HEADERS ({len(headers)} columns):")
    print(f"{'=' * 80}")
    for h in headers:
        print(f"  {h['col']:4s} (#{h['index']:3d}): {h['name']}")
    
    # Check for data validation
    print(f"\n{'=' * 80}")
    print("DROPDOWNS:")
    print(f"{'=' * 80}")
    
    for dv in ws.data_validations.dataValidation:
        if dv.type == 'list':
            formula = dv.formula1
            cells = str(dv.sqref)
            
            # Find which column this applies to
            col_letter = cells.split(':')[0].lstrip('$').rstrip('0123456789')
            matching_header = next((h for h in headers if h['col'] == col_letter), None)
            
            print(f"\n  Field: {matching_header['name'] if matching_header else col_letter}")
            print(f"  Range: {cells}")
            
            if formula:
                if formula.startswith('"'):
                    values = [v.strip() for v in formula.strip('"').split(',')]
                    print(f"  Options: {values}")
                else:
                    print(f"  Ref: {formula}")
    
    # Sample data rows to understand field types
    print(f"\n{'=' * 80}")
    print("SAMPLE DATA (first 3 rows):")
    print(f"{'=' * 80}")
    
    for row_idx in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
        print(f"\nRow {row_idx}:")
        for h in headers[:30]:  # First 30 columns
            cell = ws.cell(row_idx, h['index'])
            value = cell.value
            if value:
                print(f"  {h['name'][:40]:40s}: {str(value)[:60]}")
    
    # Export clean structure
    output = {
        'sheet_name': sheet_name,
        'header_row': header_row,
        'headers': [{'name': h['name'], 'col': h['col'], 'index': h['index']} for h in headers]
    }
    
    with open('/Users/Erin/saas-crm/door-production-structure.json', 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"\n{'=' * 80}")
    print("Exported to: door-production-structure.json")
    print(f"{'=' * 80}")

wb.close()
